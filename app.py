import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import hashlib
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ========================================================================
# CONFIGURACAO
# ========================================================================
st.set_page_config(
    page_title="Dashboard de Pedidos - Autron",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded"
)

DATA_DIR = os.environ.get("DATA_DIR", os.path.join(os.path.dirname(__file__), "dados"))

# ========================================================================
# AUTENTICACAO
# ========================================================================
# Usuarios configurados via variavel de ambiente AUTH_USERS
# Formato: "email1:senha1,email2:senha2"
# Exemplo: "dani@autron.com:Autron2026,admin@autron.com:Admin123"
# Se AUTH_USERS nao estiver definida, login fica desabilitado (acesso livre)

AUTH_USERS_RAW = os.environ.get("AUTH_USERS", "")


def get_usuarios():
    """Carrega usuarios da variavel de ambiente."""
    if not AUTH_USERS_RAW.strip():
        return {}
    usuarios = {}
    for par in AUTH_USERS_RAW.split(","):
        par = par.strip()
        if ":" in par:
            email, senha = par.split(":", 1)
            usuarios[email.strip().lower()] = senha.strip()
    return usuarios


def hash_senha(senha):
    return hashlib.sha256(senha.encode()).hexdigest()


def verificar_login(email, senha):
    usuarios = get_usuarios()
    email = email.strip().lower()
    if email in usuarios:
        return usuarios[email] == senha
    return False


def tela_login():
    """Exibe tela de login estilizada."""
    st.markdown("""
    <style>
        .login-container {
            max-width: 420px;
            margin: 80px auto;
            padding: 40px;
            background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
            border: 1px solid #2F5496;
            border-radius: 16px;
            box-shadow: 0 8px 32px rgba(0,0,0,0.3);
        }
        .login-title {
            text-align: center;
            color: #4FC3F7;
            font-size: 1.8rem;
            font-weight: 700;
            margin-bottom: 5px;
        }
        .login-subtitle {
            text-align: center;
            color: #B0BEC5;
            font-size: 0.95rem;
            margin-bottom: 30px;
        }
        .login-logo {
            text-align: center;
            font-size: 3rem;
            margin-bottom: 10px;
        }
        [data-testid="stForm"] {
            border: none !important;
        }
    </style>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 1.2, 1])
    with col2:
        st.markdown('<div class="login-logo">📦</div>', unsafe_allow_html=True)
        st.markdown('<div class="login-title">Autron</div>', unsafe_allow_html=True)
        st.markdown('<div class="login-subtitle">Dashboard de Pedidos</div>', unsafe_allow_html=True)

        with st.form("login_form"):
            email = st.text_input("📧 Email", placeholder="seu@email.com")
            senha = st.text_input("🔒 Senha", type="password", placeholder="Digite sua senha")
            submit = st.form_submit_button("Entrar", use_container_width=True, type="primary")

            if submit:
                if email and senha:
                    if verificar_login(email, senha):
                        st.session_state["autenticado"] = True
                        st.session_state["usuario"] = email.strip().lower()
                        st.rerun()
                    else:
                        st.error("Email ou senha incorretos.")
                else:
                    st.warning("Preencha email e senha.")

    st.stop()


# Verificar se login esta habilitado
LOGIN_HABILITADO = bool(AUTH_USERS_RAW.strip())

if LOGIN_HABILITADO:
    if not st.session_state.get("autenticado", False):
        tela_login()

CORES = {
    "azul": "#2F5496",
    "azul_claro": "#4472C4",
    "verde": "#2ECC71",
    "amarelo": "#F39C12",
    "vermelho": "#E74C3C",
    "vermelho_escuro": "#C0392B",
    "cinza": "#95A5A6",
    "branco": "#FFFFFF",
    "fundo": "#0E1117",
}


# ========================================================================
# CSS CUSTOMIZADO
# ========================================================================
st.markdown("""
<style>
    .main .block-container { padding-top: 1rem; max-width: 100%; }

    .kpi-card {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
        border: 1px solid #2F5496;
        border-radius: 12px;
        padding: 20px;
        text-align: center;
        margin: 5px 0;
    }
    .kpi-value {
        font-size: 2.2rem;
        font-weight: 700;
        color: #4FC3F7;
        margin: 0;
    }
    .kpi-label {
        font-size: 0.85rem;
        color: #B0BEC5;
        margin-top: 5px;
    }
    .kpi-card-alert {
        background: linear-gradient(135deg, #4a1a1a 0%, #3e1616 100%);
        border: 1px solid #E74C3C;
        border-radius: 12px;
        padding: 20px;
        text-align: center;
        margin: 5px 0;
    }
    .kpi-value-alert {
        font-size: 2.2rem;
        font-weight: 700;
        color: #E74C3C;
        margin: 0;
    }
    .kpi-card-ok {
        background: linear-gradient(135deg, #1a3a1a 0%, #163e16 100%);
        border: 1px solid #2ECC71;
        border-radius: 12px;
        padding: 20px;
        text-align: center;
        margin: 5px 0;
    }
    .kpi-value-ok {
        font-size: 2.2rem;
        font-weight: 700;
        color: #2ECC71;
        margin: 0;
    }
    .kpi-card-warn {
        background: linear-gradient(135deg, #3a3a1a 0%, #3e3616 100%);
        border: 1px solid #F39C12;
        border-radius: 12px;
        padding: 20px;
        text-align: center;
        margin: 5px 0;
    }
    .kpi-value-warn {
        font-size: 2.2rem;
        font-weight: 700;
        color: #F39C12;
        margin: 0;
    }

    div[data-testid="stMetric"] {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
        border: 1px solid #2F5496;
        border-radius: 10px;
        padding: 15px;
    }

    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        background-color: #1a1a2e;
        border-radius: 8px 8px 0 0;
        padding: 10px 20px;
        border: 1px solid #2F5496;
    }
    .stTabs [aria-selected="true"] {
        background-color: #2F5496;
    }

    h1, h2, h3 { color: #E0E0E0; }

    .status-finalizado { color: #2ECC71; font-weight: bold; }
    .status-aberto { color: #E74C3C; font-weight: bold; }

    .upload-section {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
        border: 2px dashed #2F5496;
        border-radius: 12px;
        padding: 20px;
        margin: 10px 0;
    }
</style>
""", unsafe_allow_html=True)


def kpi_card(valor, label, tipo="normal"):
    css_class = {
        "normal": ("kpi-card", "kpi-value"),
        "alert": ("kpi-card-alert", "kpi-value-alert"),
        "ok": ("kpi-card-ok", "kpi-value-ok"),
        "warn": ("kpi-card-warn", "kpi-value-warn"),
    }.get(tipo, ("kpi-card", "kpi-value"))

    return f"""
    <div class="{css_class[0]}">
        <p class="{css_class[1]}">{valor}</p>
        <p class="kpi-label">{label}</p>
    </div>
    """


# ========================================================================
# CARGA E PROCESSAMENTO DE DADOS
# ========================================================================
@st.cache_data(ttl=300)
def carregar_e_processar():
    """Carrega os arquivos e processa toda a logica de negocio."""

    arq_entrada = os.path.join(DATA_DIR, "entrada_pedido.xlsx")
    arq_followup = os.path.join(DATA_DIR, "followup.xlsx")
    arq_estoque = os.path.join(DATA_DIR, "matr260.xlsx")
    arq_faturamento = os.path.join(DATA_DIR, "faturamento.xlsx")

    # CSV pode ter nomes diferentes
    arq_scio = os.path.join(DATA_DIR, "sciozvs0.csv")
    if not os.path.exists(arq_scio):
        arq_scio = os.path.join(DATA_DIR, "sciozmq0.csv")

    obrigatorios = [arq_entrada, arq_followup, arq_estoque, arq_scio]
    for arq in obrigatorios:
        if not os.path.exists(arq):
            return None, None, f"Arquivo nao encontrado: {os.path.basename(arq)}"

    # Carregar
    ep = pd.read_excel(arq_entrada, header=1)
    fu = pd.read_excel(arq_followup, header=1)

    # MATR260 - header fixo na linha 6
    mt = pd.read_excel(arq_estoque, header=6)

    sc_csv = pd.read_csv(arq_scio, encoding='latin-1', sep=';', header=2, low_memory=False)

    # Faturamento (opcional)
    fat = None
    if os.path.exists(arq_faturamento):
        fat = pd.read_excel(arq_faturamento, header=1)

    # Limpar entrada_pedido
    ep['Num. Pedido'] = ep['Num. Pedido'].astype(str).str.replace('.', '', regex=False).str.strip()
    ep['Numero SC'] = pd.to_numeric(ep['Numero SC'], errors='coerce')
    ep['Quantidade'] = pd.to_numeric(ep['Quantidade'], errors='coerce')
    ep['DT Emissao'] = pd.to_datetime(ep['DT Emissao'], errors='coerce')
    ep['DT. Ofertada'] = pd.to_datetime(ep['DT. Ofertada'], errors='coerce')
    ep['DT. Fat. Cli'] = pd.to_datetime(ep['DT. Fat. Cli'], errors='coerce')
    ep['Nota Fiscal'] = pd.to_numeric(ep['Nota Fiscal'], errors='coerce')
    ep['Item'] = pd.to_numeric(ep['Item'], errors='coerce')
    ep['Vlr.Total'] = pd.to_numeric(ep['Vlr.Total'], errors='coerce')
    ep['Margem'] = pd.to_numeric(ep['Margem'], errors='coerce')
    if 'Numero OP' not in ep.columns:
        ep['Numero OP'] = np.nan

    # Filtrar PVs que iniciam com I ou B (desconsiderar completamente)
    ep = ep[~ep['Num. Pedido'].str.upper().str.match(r'^[IB]', na=False)]

    # Limpar follow-up
    fu['No. da S.C.'] = pd.to_numeric(fu['No. da S.C.'], errors='coerce')
    fu['Dt. Confirma'] = pd.to_datetime(fu['Dt. Confirma'], errors='coerce')
    fu['Dt. Pre entr'] = pd.to_datetime(fu['Dt. Pre entr'], errors='coerce')
    fu['Dt Chegada Autron'] = pd.to_datetime(fu['Dt Chegada Autron'], errors='coerce')
    fu['PV informado na SC'] = pd.to_numeric(fu['PV informado na SC'], errors='coerce')
    fu['Numero do PV'] = fu['Numero do PV'].astype(str).str.replace('.', '', regex=False).str.strip()
    if 'OP na SC' not in fu.columns:
        fu['OP na SC'] = np.nan

    # Limpar estoque (MATR260)
    mt['SALDO EM ESTOQUE'] = pd.to_numeric(mt['SALDO EM ESTOQUE'], errors='coerce').fillna(0)

    # Classificacao Comprando/Produzindo
    sc_csv['Produto'] = sc_csv['Produto'].astype(str).str.strip()
    sc_csv['Prod.Solict'] = sc_csv['Prod.Solict'].astype(str).str.strip()
    tipo_prod = sc_csv.groupby('Produto')['Prod.Solict'].agg(
        lambda x: x.mode().iloc[0] if len(x.mode()) > 0 else 'Indefinido'
    ).reset_index()
    tipo_prod.rename(columns={'Prod.Solict': 'Tipo_Produto'}, inplace=True)

    # --- Follow-up por SC ---
    fu_por_sc = fu.groupby('No. da S.C.').agg({
        'Dt. Confirma': 'first', 'Dt. Pre entr': 'first', 'Pasta': 'first',
        'No. do P.O.': 'first', 'Dt Chegada Autron': 'first', 'OP na SC': 'first'
    }).reset_index()
    fu_por_sc.rename(columns={'No. da S.C.': 'Numero SC'}, inplace=True)

    # --- Follow-up por PV+Produto ---
    fu_pv = fu.copy()
    fu_pv['Numero do PV'] = fu_pv['Numero do PV'].str.strip()
    fu_por_pv = fu_pv.groupby(['Numero do PV', 'Codigo Item']).agg({
        'Dt. Confirma': 'first', 'Dt. Pre entr': 'first', 'Pasta': 'first',
        'No. do P.O.': 'first', 'Dt Chegada Autron': 'first', 'OP na SC': 'first'
    }).reset_index()
    fu_por_pv.rename(columns={
        'Numero do PV': 'Num. Pedido', 'Codigo Item': 'Produto',
        'Dt. Confirma': 'FU_Dt_Confirma_PV', 'Dt. Pre entr': 'FU_Dt_Pre_entr_PV',
        'Pasta': 'FU_Pasta_PV', 'No. do P.O.': 'FU_PO_PV',
        'Dt Chegada Autron': 'FU_Dt_Chegada_PV', 'OP na SC': 'FU_OP_na_SC_PV'
    }, inplace=True)

    # Merges
    merged = ep.merge(fu_por_sc, on='Numero SC', how='left', suffixes=('', '_fu'))
    merged = merged.merge(fu_por_pv, on=['Num. Pedido', 'Produto'], how='left')

    # Consolidar
    merged['FU_Dt_Confirma'] = merged['Dt. Confirma'].combine_first(merged['FU_Dt_Confirma_PV'])
    merged['FU_Dt_Pre_Entr'] = merged['Dt. Pre entr'].combine_first(merged['FU_Dt_Pre_entr_PV'])
    merged['FU_Pasta'] = merged['Pasta'].combine_first(merged['FU_Pasta_PV'])
    merged['FU_PO'] = merged['No. do P.O.'].combine_first(merged['FU_PO_PV'])
    merged['FU_Dt_Chegada_Autron'] = merged['Dt Chegada Autron'].combine_first(merged['FU_Dt_Chegada_PV'])
    merged['FU_OP_na_SC'] = merged['OP na SC'].combine_first(merged['FU_OP_na_SC_PV'])

    merged['Prazo_Real_Entrega'] = merged['FU_Dt_Confirma'].combine_first(merged['FU_Dt_Pre_Entr']).astype(object)
    merged['Semana_Entrega'] = np.where(merged['FU_Dt_Confirma'].notna(), merged['FU_Pasta'], None)

    # Tipo produto
    merged = merged.merge(tipo_prod, on='Produto', how='left')
    merged['Tipo_Produto'] = merged['Tipo_Produto'].fillna('Nao classificado')

    # Status
    merged['Status_Pedido'] = np.where(merged['Nota Fiscal'].notna(), 'FINALIZADO', 'EM ABERTO')

    # Estoque (MATR260)
    stock = mt.groupby('CODIGO')['SALDO EM ESTOQUE'].sum().reset_index()
    stock.rename(columns={'CODIGO': 'Produto', 'SALDO EM ESTOQUE': 'Estoque_Disponivel'}, inplace=True)
    merged = merged.merge(stock, on='Produto', how='left')
    merged['Estoque_Disponivel'] = merged['Estoque_Disponivel'].fillna(0)

    # Alocacao
    open_orders = merged[merged['Status_Pedido'] == 'EM ABERTO'].copy()
    open_orders = open_orders.sort_values(['Produto', 'DT Emissao', 'Num. Pedido', 'Item'])

    allocation = {}
    for produto in open_orders['Produto'].unique():
        prod_orders = open_orders[open_orders['Produto'] == produto]
        available = prod_orders['Estoque_Disponivel'].iloc[0] if len(prod_orders) > 0 else 0
        for idx, row in prod_orders.iterrows():
            qty = row['Quantidade'] if pd.notna(row['Quantidade']) else 0
            alloc = min(available, qty)
            allocation[idx] = {
                'Qtd_Alocada': alloc,
                'Disponivel_Estoque': 'SIM' if alloc >= qty else ('PARCIAL' if alloc > 0 else 'NAO')
            }
            available = max(0, available - qty)

    alloc_df = pd.DataFrame.from_dict(allocation, orient='index')
    merged = merged.join(alloc_df, how='left')
    merged.loc[merged['Status_Pedido'] == 'FINALIZADO', 'Disponivel_Estoque'] = 'N/A'
    merged.loc[merged['Status_Pedido'] == 'FINALIZADO', 'Qtd_Alocada'] = None

    # Identificar itens de servico pela descricao (contem "servico" ou "serviço")
    merged['_eh_servico'] = merged['Descricao do Produto'].astype(str).str.lower().str.contains(
        r'servi[cç]o', regex=True, na=False
    )
    # Para servicos: disponibilidade = "Serviço", sem alocacao de estoque
    merged.loc[merged['_eh_servico'] & (merged['Status_Pedido'] == 'EM ABERTO'), 'Disponivel_Estoque'] = 'Serviço'
    merged.loc[merged['_eh_servico'] & (merged['Status_Pedido'] == 'EM ABERTO'), 'Qtd_Alocada'] = None

    # IND21 com "posto" ou "cabine" na descricao: prazo "A definir"
    _eh_ind21_posto_cabine = (
        merged['Unidade Negocio'].astype(str).str.strip().str.upper().eq('IND21') &
        merged['Descricao do Produto'].astype(str).str.lower().str.contains(r'posto|cabine', regex=True, na=False) &
        (merged['Status_Pedido'] == 'EM ABERTO')
    )
    merged.loc[_eh_ind21_posto_cabine, 'Prazo_Real_Entrega'] = 'A definir'
    merged.loc[_eh_ind21_posto_cabine, 'FU_Dt_Confirma'] = None
    merged.loc[_eh_ind21_posto_cabine, 'FU_Dt_Pre_Entr'] = None

    # Regras SC/OP
    merged['Tem_SC'] = merged['Numero SC'].notna()
    merged['Tem_OP'] = (
        (merged['Numero OP'].notna() & merged['Numero OP'].astype(str).str.strip().ne('') & merged['Numero OP'].astype(str).str.strip().ne('nan')) |
        (merged['FU_OP_na_SC'].notna() & merged['FU_OP_na_SC'].astype(str).str.strip().ne('') & merged['FU_OP_na_SC'].astype(str).str.strip().ne('nan'))
    )

    def gerar_acao(row):
        if row['Status_Pedido'] == 'FINALIZADO':
            return 'Finalizado'
        if row.get('_eh_servico', False):
            return 'Prazo a confirmar'
        if row['Disponivel_Estoque'] == 'SIM':
            return 'Estoque OK'
        tipo = str(row.get('Tipo_Produto', '')).strip()
        if tipo == 'Comprando':
            if row['Tem_OP']: return 'ERRO no CADASTRO'
            elif row['Tem_SC']: return 'SC gerada - Aguardando'
            else: return 'Necessario gerar SC'
        elif tipo == 'Produzindo':
            if row['Tem_OP']: return 'OP gerada - Aguardando'
            else: return 'Necessario gerar OP'
        else:
            return 'Verificar classificacao'

    merged['Acao_Necessaria'] = merged.apply(gerar_acao, axis=1)

    # Indicadores
    today = pd.Timestamp.now().normalize()
    merged['Dias_Atraso_Cliente'] = np.where(
        (merged['Status_Pedido'] == 'EM ABERTO') & (merged['DT. Fat. Cli'].notna()),
        (today - merged['DT. Fat. Cli']).dt.days, None
    )
    merged['Dias_Atraso_Cliente'] = pd.to_numeric(merged['Dias_Atraso_Cliente'], errors='coerce')

    merged['Dias_Atraso_Ofertada'] = np.where(
        (merged['Status_Pedido'] == 'EM ABERTO') & (merged['DT. Ofertada'].notna()),
        (today - merged['DT. Ofertada']).dt.days, None
    )
    merged['Dias_Atraso_Ofertada'] = pd.to_numeric(merged['Dias_Atraso_Ofertada'], errors='coerce')

    merged['Mes_Emissao'] = merged['DT Emissao'].dt.to_period('M').astype(str)
    merged['Ano_Emissao'] = merged['DT Emissao'].dt.year

    merged['Pronto_para_Fazer'] = np.where(
        merged['Status_Pedido'] == 'FINALIZADO', 'FINALIZADO',
        np.where(
            (merged['Disponivel_Estoque'] == 'SIM') & (merged['FU_Dt_Confirma'].notna()), 'SIM',
            np.where(
                merged['Disponivel_Estoque'] == 'SIM', 'PARCIAL - Sem Follow-up',
                np.where(
                    merged['FU_Dt_Confirma'].notna(), 'PARCIAL - Sem Estoque', 'NAO'
                )
            )
        )
    )

    return merged, fat, None


# ========================================================================
# GERAR EXCEL PARA DOWNLOAD
# ========================================================================
def gerar_excel_consolidado(df, fat=None):
    """Gera o Excel consolidado no mesmo formato do dashboard_pedidos_powerbi.xlsx."""
    wb = Workbook()
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor='2F5496')
    data_font = Font(name='Arial', size=9)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )
    fill_verde = PatternFill('solid', fgColor='C6EFCE')
    fill_amarelo = PatternFill('solid', fgColor='FFEB9C')
    fill_vermelho = PatternFill('solid', fgColor='FFC7CE')
    fill_erro = PatternFill('solid', fgColor='FF4444')
    font_erro = Font(name='Arial', size=9, bold=True, color='FFFFFF')

    output_cols = [
        'Num. Pedido', 'Item', 'Cliente', 'Razao Social', 'Nome Fantasia',
        'Produto', 'Descricao do Produto', 'Tipo_Produto', 'Quantidade',
        'Prc Unitario', 'Vlr.Total', 'Margem',
        'DT Emissao', 'Mes_Emissao', 'Ano_Emissao',
        'DT. Ofertada', 'DT. Fat. Cli', 'Ped Cliente',
        'Prazo_Real_Entrega', 'FU_Dt_Confirma', 'FU_Dt_Pre_Entr', 'Semana_Entrega',
        'FU_Dt_Chegada_Autron', 'FU_PO', 'FU_OP_na_SC',
        'Nota Fiscal', 'Status_Pedido',
        'Estoque_Disponivel', 'Qtd_Alocada', 'Disponivel_Estoque',
        'Acao_Necessaria', 'Pronto_para_Fazer',
        'Dias_Atraso_Ofertada', 'Dias_Atraso_Cliente',
        'Numero SC', 'Numero OP', 'TP Venda (PV)', 'Tipo Negocio (PV)',
        'Nome do Vendedor', 'Familia', 'Fabricante',
        'Unidade Negocio', 'Estado', 'Regional (PV)',
        'Nome do Segmento 1', 'Nome do Segmento 2', 'Nome do Segmento 3'
    ]
    header_map = {
        'Num. Pedido': 'PV', 'Item': 'Item', 'Cliente': 'Cod Cliente',
        'Razao Social': 'Razao Social', 'Nome Fantasia': 'Nome Fantasia',
        'Produto': 'Codigo Produto', 'Descricao do Produto': 'Descricao Produto',
        'Tipo_Produto': 'Tipo (Comprando/Produzindo)',
        'Quantidade': 'Qtd', 'Prc Unitario': 'Preco Unit.', 'Vlr.Total': 'Valor Total',
        'Margem': 'Margem %', 'DT Emissao': 'Data Emissao', 'Mes_Emissao': 'Mes Emissao',
        'Ano_Emissao': 'Ano Emissao', 'DT. Ofertada': 'Data Ofertada',
        'DT. Fat. Cli': 'Data Solicitada Cliente', 'Ped Cliente': 'Pedido Cliente',
        'Prazo_Real_Entrega': 'Prazo Real Entrega', 'FU_Dt_Confirma': 'Follow-Up Dt Confirmada',
        'FU_Dt_Pre_Entr': 'Follow-Up Dt Pre-Entrega', 'Semana_Entrega': 'Semana Entrega (Pasta)',
        'FU_Dt_Chegada_Autron': 'Dt Chegada Autron',
        'FU_PO': 'Purchase Order', 'FU_OP_na_SC': 'OP na SC',
        'Nota Fiscal': 'Nota Fiscal', 'Status_Pedido': 'Status',
        'Estoque_Disponivel': 'Estoque Disponivel', 'Qtd_Alocada': 'Qtd Alocada Estoque',
        'Disponivel_Estoque': 'Disponibilidade Estoque',
        'Acao_Necessaria': 'Acao Necessaria', 'Pronto_para_Fazer': 'Pronto p/ Fazer?',
        'Dias_Atraso_Ofertada': 'Dias Atraso vs Ofertada',
        'Dias_Atraso_Cliente': 'Dias Atraso vs Cliente',
        'Numero SC': 'No SC', 'Numero OP': 'No OP',
        'TP Venda (PV)': 'Tipo Venda', 'Tipo Negocio (PV)': 'Tipo Negocio',
        'Nome do Vendedor': 'Vendedor', 'Familia': 'Familia', 'Fabricante': 'Fabricante',
        'Unidade Negocio': 'Unidade Negocio', 'Estado': 'Estado',
        'Regional (PV)': 'Regional', 'Nome do Segmento 1': 'Segmento 1',
        'Nome do Segmento 2': 'Segmento 2', 'Nome do Segmento 3': 'Segmento 3'
    }

    cols = [c for c in output_cols if c in df.columns]
    headers = [header_map.get(c, c) for c in cols]
    output = df[cols].copy().sort_values(['Status_Pedido', 'DT Emissao', 'Num. Pedido', 'Item'])

    date_cols = {'DT Emissao', 'DT. Ofertada', 'DT. Fat. Cli',
                 'FU_Dt_Confirma', 'FU_Dt_Pre_Entr', 'FU_Dt_Chegada_Autron'}
    # Prazo_Real_Entrega pode ser data ou string ("A definir"), tratar separadamente
    date_or_str_cols = {'Prazo_Real_Entrega'}
    col_idx_map = {c: i for i, c in enumerate(cols)}

    # --- Aba Pedidos_Consolidado ---
    ws = wb.active
    ws.title = 'Pedidos_Consolidado'
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for ri, (_, row) in enumerate(output.iterrows(), 2):
        for ci, col_name in enumerate(cols, 1):
            val = row[col_name]
            if pd.isna(val): val = None
            elif isinstance(val, pd.Timestamp): val = val.to_pydatetime()
            elif isinstance(val, (np.int64, np.float64)):
                val = float(val) if not np.isnan(val) else None
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = data_font
            cell.border = thin_border
            if col_name in date_cols and val is not None: cell.number_format = 'DD/MM/YYYY'
            elif col_name in date_or_str_cols and isinstance(val, pd.Timestamp): cell.number_format = 'DD/MM/YYYY'
            elif col_name == 'Margem' and val is not None: cell.number_format = '0.00%'
            elif col_name in ('Prc Unitario', 'Vlr.Total') and val is not None: cell.number_format = '#,##0.00'

        # Cores condicionais
        for field, fills in [
            ('Status_Pedido', {'FINALIZADO': fill_verde, 'EM ABERTO': fill_vermelho}),
            ('Disponivel_Estoque', {'SIM': fill_verde, 'PARCIAL': fill_amarelo, 'NAO': fill_vermelho, 'Serviço': fill_amarelo}),
        ]:
            idx = col_idx_map.get(field)
            if idx is not None:
                c = ws.cell(row=ri, column=idx + 1)
                f = fills.get(c.value)
                if f: c.fill = f

        idx = col_idx_map.get('Pronto_para_Fazer')
        if idx is not None:
            c = ws.cell(row=ri, column=idx + 1)
            v = c.value
            if v in ('SIM', 'FINALIZADO'): c.fill = fill_verde
            elif v and 'PARCIAL' in str(v): c.fill = fill_amarelo
            elif v == 'NAO': c.fill = fill_vermelho

        idx = col_idx_map.get('Acao_Necessaria')
        if idx is not None:
            c = ws.cell(row=ri, column=idx + 1)
            v = c.value
            if v == 'ERRO no CADASTRO': c.fill = fill_erro; c.font = font_erro
            elif v == 'Estoque OK': c.fill = fill_verde
            elif v == 'Prazo a confirmar': c.fill = fill_amarelo
            elif v and 'Necessario' in str(v): c.fill = fill_vermelho
            elif v and ('gerada' in str(v) or 'Aguardando' in str(v)): c.fill = fill_amarelo

    for ci in range(1, len(headers) + 1):
        max_len = len(str(ws.cell(row=1, column=ci).value or ''))
        for ri in range(2, min(100, ws.max_row + 1)):
            max_len = max(max_len, len(str(ws.cell(row=ri, column=ci).value or '')))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 35)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{ws.max_row}'

    # --- Aba Estoque_Pedidos_Abertos ---
    ws2 = wb.create_sheet('Estoque_Pedidos_Abertos')
    open_cols = [c for c in ['Produto', 'Descricao do Produto', 'Tipo_Produto', 'Num. Pedido', 'Item',
        'Quantidade', 'Estoque_Disponivel', 'Qtd_Alocada', 'Disponivel_Estoque',
        'Acao_Necessaria', 'Numero SC', 'Numero OP', 'FU_OP_na_SC', 'DT Emissao'] if c in output.columns]
    open_data = output[output['Status_Pedido'] == 'EM ABERTO'][open_cols].sort_values(['Produto', 'DT Emissao'])
    headers2 = ['Codigo Produto', 'Descricao', 'Tipo (Comp/Prod)', 'PV', 'Item', 'Qtd Pedida',
                'Estoque Total', 'Qtd Alocada', 'Disponibilidade', 'Acao Necessaria',
                'No SC', 'No OP', 'OP na SC (Follow-up)', 'Data Emissao PV']
    for ci, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    for ri, (_, row) in enumerate(open_data.iterrows(), 2):
        for ci, col in enumerate(open_data.columns, 1):
            val = row[col]
            if pd.isna(val): val = None
            elif isinstance(val, pd.Timestamp): val = val.to_pydatetime()
            elif isinstance(val, (np.int64, np.float64)):
                val = float(val) if not np.isnan(val) else None
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.font = data_font; cell.border = thin_border
            if col == 'DT Emissao' and val: cell.number_format = 'DD/MM/YYYY'
        c = ws2.cell(row=ri, column=9)
        if c.value == 'SIM': c.fill = fill_verde
        elif c.value == 'PARCIAL': c.fill = fill_amarelo
        elif c.value == 'NAO': c.fill = fill_vermelho
        elif c.value == 'Serviço': c.fill = fill_amarelo
        c2 = ws2.cell(row=ri, column=10)
        if c2.value == 'ERRO no CADASTRO': c2.fill = fill_erro; c2.font = font_erro
        elif c2.value == 'Estoque OK': c2.fill = fill_verde
        elif c2.value == 'Prazo a confirmar': c2.fill = fill_amarelo
        elif c2.value and 'Necessario' in str(c2.value): c2.fill = fill_vermelho
        elif c2.value and 'gerada' in str(c2.value): c2.fill = fill_amarelo
    for ci in range(1, len(headers2) + 1):
        max_len = len(str(ws2.cell(row=1, column=ci).value or ''))
        for ri in range(2, min(50, ws2.max_row + 1)):
            max_len = max(max_len, len(str(ws2.cell(row=ri, column=ci).value or '')))
        ws2.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 40)
    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f'A1:{get_column_letter(len(headers2))}{ws2.max_row}'

    # --- Aba Faturamento (se disponivel) ---
    if fat is not None and len(fat) > 0:
        ws3 = wb.create_sheet('Faturamento')
        fat_cols = [c for c in ['Emissao', 'Num. Docto.', 'Serie', 'No do Pedido', 'Item Pv', 'Produto',
            'Descricao Produto', 'Quantidade', 'Vlr.Unitario', 'Vlr. Total (quant * preco un)',
            'Vlr.Bruto', 'Cliente', 'Razao Social', 'Nome Fantasia', 'UF',
            'Faturamento Bruto', 'Faturamento Liquido', 'Margem Liquida (R$)',
            'Margem Liquida (%) por NF Faturada', 'Nome do Vendedor', 'Tipo Negocio',
            'TP Venda', 'Regional', 'Unid Negocio'] if c in fat.columns]
        for ci, h in enumerate(fat_cols, 1):
            cell = ws3.cell(row=1, column=ci, value=h)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        for ri, (_, row) in enumerate(fat[fat_cols].iterrows(), 2):
            for ci, col in enumerate(fat_cols, 1):
                val = row[col]
                if pd.isna(val): val = None
                elif isinstance(val, pd.Timestamp): val = val.to_pydatetime()
                elif isinstance(val, (np.int64, np.float64)):
                    val = float(val) if not np.isnan(val) else None
                cell = ws3.cell(row=ri, column=ci, value=val)
                cell.font = data_font; cell.border = thin_border
                if col == 'Emissao' and val: cell.number_format = 'DD/MM/YYYY'
                elif col in ('Vlr.Unitario', 'Vlr. Total (quant * preco un)', 'Vlr.Bruto',
                              'Faturamento Bruto', 'Faturamento Liquido', 'Margem Liquida (R$)'):
                    if val is not None: cell.number_format = '#,##0.00'
        for ci in range(1, len(fat_cols) + 1):
            max_len = len(str(ws3.cell(row=1, column=ci).value or ''))
            for ri in range(2, min(50, ws3.max_row + 1)):
                max_len = max(max_len, len(str(ws3.cell(row=ri, column=ci).value or '')))
            ws3.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 35)
        ws3.freeze_panes = 'A2'
        ws3.auto_filter.ref = f'A1:{get_column_letter(len(fat_cols))}{ws3.max_row}'

    # Salvar em buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ========================================================================
# UPLOAD E GESTAO DE ARQUIVOS
# ========================================================================
os.makedirs(DATA_DIR, exist_ok=True)
ARQUIVOS_OBRIGATORIOS = ['entrada_pedido.xlsx', 'followup.xlsx', 'matr260.xlsx']
ARQUIVOS_CSV = ['sciozvs0.csv', 'sciozmq0.csv']  # aceita qualquer um dos dois
ARQUIVO_OPCIONAL = 'faturamento.xlsx'


def verificar_arquivos():
    """Retorna lista de arquivos presentes e ausentes."""
    presentes = []
    ausentes = []
    for f in ARQUIVOS_OBRIGATORIOS:
        caminho = os.path.join(DATA_DIR, f)
        if os.path.exists(caminho) and os.path.getsize(caminho) > 0:
            mod_time = datetime.fromtimestamp(os.path.getmtime(caminho))
            presentes.append((f, mod_time))
        else:
            ausentes.append(f)
    # CSV - qualquer um dos dois nomes
    csv_encontrado = False
    for csv_name in ARQUIVOS_CSV:
        caminho = os.path.join(DATA_DIR, csv_name)
        if os.path.exists(caminho) and os.path.getsize(caminho) > 0:
            mod_time = datetime.fromtimestamp(os.path.getmtime(caminho))
            presentes.append((csv_name, mod_time))
            csv_encontrado = True
            break
    if not csv_encontrado:
        ausentes.append('sciozvs0.csv (ou sciozmq0.csv)')
    # Faturamento (opcional)
    caminho = os.path.join(DATA_DIR, ARQUIVO_OPCIONAL)
    if os.path.exists(caminho) and os.path.getsize(caminho) > 0:
        mod_time = datetime.fromtimestamp(os.path.getmtime(caminho))
        presentes.append((ARQUIVO_OPCIONAL + ' (opcional)', mod_time))
    return presentes, ausentes


def tela_upload():
    """Exibe tela de upload dos arquivos."""
    st.markdown("""
    <div style="text-align: center; padding: 40px 0 20px 0;">
        <h1 style="color: #4FC3F7;">📦 Dashboard de Pedidos</h1>
        <p style="color: #B0BEC5; font-size: 1.1rem;">
            Envie os relatorios para gerar o dashboard
        </p>
    </div>
    """, unsafe_allow_html=True)

    presentes, ausentes = verificar_arquivos()

    if presentes:
        st.markdown("#### ✅ Arquivos ja carregados:")
        for nome, dt in presentes:
            st.markdown(f"- **{nome}** — atualizado em {dt.strftime('%d/%m/%Y %H:%M')}")
        st.markdown("")

    if ausentes:
        st.markdown(f"#### 📤 Arquivos pendentes ({len(ausentes)}):")
        for nome in ausentes:
            st.markdown(f"- ❌ {nome}")
        st.markdown("")

    st.markdown("---")
    st.markdown("### Envie os arquivos:")

    col1, col2 = st.columns(2)
    with col1:
        up_entrada = st.file_uploader(
            "📋 entrada_pedido.xlsx",
            type=['xlsx'], key='up_entrada',
            help="Relatorio de entrada de pedidos do ERP"
        )
        up_followup = st.file_uploader(
            "📅 followup.xlsx",
            type=['xlsx'], key='up_followup',
            help="Relatorio de follow-up de compras"
        )
        up_faturamento = st.file_uploader(
            "💰 faturamento.xlsx (opcional)",
            type=['xlsx'], key='up_faturamento',
            help="Relatorio de faturamento"
        )
    with col2:
        up_mata = st.file_uploader(
            "📦 matr260.xlsx",
            type=['xlsx'], key='up_mata',
            help="Relatorio de estoque (MATR260)"
        )
        up_scio = st.file_uploader(
            "🏭 sciozvs0.csv / sciozmq0.csv",
            type=['csv'], key='up_scio',
            help="Relatorio SC com classificacao Comprando/Produzindo"
        )

    uploads = {
        'entrada_pedido.xlsx': up_entrada,
        'followup.xlsx': up_followup,
        'matr260.xlsx': up_mata,
    }

    # Salvar arquivos enviados
    algum_novo = False
    for nome, uploaded in uploads.items():
        if uploaded is not None:
            with open(os.path.join(DATA_DIR, nome), 'wb') as f:
                f.write(uploaded.getbuffer())
            algum_novo = True

    # CSV - salvar com o nome original do arquivo
    if up_scio is not None:
        csv_save_name = up_scio.name if up_scio.name in ARQUIVOS_CSV else 'sciozvs0.csv'
        with open(os.path.join(DATA_DIR, csv_save_name), 'wb') as f:
            f.write(up_scio.getbuffer())
        algum_novo = True

    # Faturamento (opcional)
    if up_faturamento is not None:
        with open(os.path.join(DATA_DIR, ARQUIVO_OPCIONAL), 'wb') as f:
            f.write(up_faturamento.getbuffer())
        algum_novo = True

    # Verificar novamente apos upload
    presentes, ausentes = verificar_arquivos()

    if len(ausentes) == 0:
        st.success("✅ Todos os arquivos prontos! Processando dashboard...")
        st.cache_data.clear()
        st.rerun()
    elif algum_novo:
        st.info(f"Faltam {len(ausentes)} arquivo(s): {', '.join(ausentes)}")
        st.rerun()
    else:
        st.info(f"Envie os {len(ausentes)} arquivo(s) pendente(s) para continuar.")
        st.stop()


# Verificar se precisa mostrar tela de upload
presentes, ausentes = verificar_arquivos()
if len(ausentes) > 0:
    tela_upload()


# ========================================================================
# CARREGAR DADOS
# ========================================================================
df, fat_df, erro = carregar_e_processar()

if erro:
    st.error(f"Erro ao carregar dados: {erro}")
    st.warning("Pode haver um problema com os arquivos. Tente enviar novamente.")
    if st.button("🔄 Reenviar arquivos"):
        for f in ARQUIVOS_OBRIGATORIOS + ARQUIVOS_CSV + [ARQUIVO_OPCIONAL]:
            caminho = os.path.join(DATA_DIR, f)
            if os.path.exists(caminho):
                os.remove(caminho)
        st.cache_data.clear()
        st.rerun()
    st.stop()


# ========================================================================
# SIDEBAR - FILTROS
# ========================================================================
with st.sidebar:
    logo_path = os.path.join(os.path.dirname(__file__), "assets", "logo.png")
    if os.path.exists(logo_path):
        st.image(logo_path, width=180)
    st.markdown("## 🔍 Filtros")
    st.markdown("---")

    # Status
    status_opts = sorted(df['Status_Pedido'].unique())
    status_sel = st.multiselect("Status", status_opts, default=['EM ABERTO'])

    # Ano
    anos = sorted(df['Ano_Emissao'].dropna().unique())
    ano_sel = st.multiselect("Ano Emissao", anos, default=anos[-3:] if len(anos) >= 3 else anos)

    # Mes
    meses_disp = sorted(df[df['Ano_Emissao'].isin(ano_sel)]['Mes_Emissao'].dropna().unique()) if ano_sel else []
    mes_sel = st.multiselect("Mes Emissao", meses_disp, default=[])

    # Tipo Produto
    tipo_opts = sorted(df['Tipo_Produto'].unique())
    tipo_sel = st.multiselect("Tipo Produto", tipo_opts, default=[])

    # Vendedor
    vendedores = sorted(df['Nome do Vendedor'].dropna().unique())
    vend_sel = st.multiselect("Vendedor", vendedores, default=[])

    # Pronto para Fazer
    pronto_opts = sorted(df['Pronto_para_Fazer'].unique())
    pronto_sel = st.multiselect("Pronto p/ Fazer?", pronto_opts, default=[])

    st.markdown("---")

    # Mostrar data dos arquivos carregados
    presentes, _ = verificar_arquivos()
    if presentes:
        st.markdown("##### 📁 Arquivos carregados")
        for nome, dt in presentes:
            st.caption(f"{nome}: {dt.strftime('%d/%m/%Y %H:%M')}")

    # Botao para atualizar arquivos
    if st.button("📤 Atualizar Arquivos", use_container_width=True):
        for f in ARQUIVOS_OBRIGATORIOS + ARQUIVOS_CSV + [ARQUIVO_OPCIONAL]:
            caminho = os.path.join(DATA_DIR, f)
            if os.path.exists(caminho):
                os.remove(caminho)
        st.cache_data.clear()
        st.rerun()

    st.markdown("---")

    # Botao download Excel
    st.markdown("##### 📥 Exportar Planilha")
    if st.button("Gerar Excel Consolidado", use_container_width=True, type="primary"):
        with st.spinner("Gerando planilha..."):
            excel_buffer = gerar_excel_consolidado(df, fat_df)
            st.session_state['excel_pronto'] = excel_buffer

    if 'excel_pronto' in st.session_state:
        st.download_button(
            label="⬇️ Baixar Planilha",
            data=st.session_state['excel_pronto'],
            file_name=f"dashboard_pedidos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    st.markdown("---")
    ultima_att = datetime.now().strftime('%d/%m/%Y %H:%M')
    st.caption(f"Ultima atualizacao: {ultima_att}")

    if st.button("🔄 Atualizar Dados", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

    # Usuario logado e logout
    if LOGIN_HABILITADO:
        st.markdown("---")
        usuario = st.session_state.get("usuario", "")
        st.caption(f"👤 {usuario}")
        if st.button("🚪 Sair", use_container_width=True):
            st.session_state["autenticado"] = False
            st.session_state["usuario"] = ""
            st.rerun()

# Aplicar filtros
mask = pd.Series(True, index=df.index)
if status_sel:
    mask &= df['Status_Pedido'].isin(status_sel)
if ano_sel:
    mask &= df['Ano_Emissao'].isin(ano_sel)
if mes_sel:
    mask &= df['Mes_Emissao'].isin(mes_sel)
if tipo_sel:
    mask &= df['Tipo_Produto'].isin(tipo_sel)
if vend_sel:
    mask &= df['Nome do Vendedor'].isin(vend_sel)
if pronto_sel:
    mask &= df['Pronto_para_Fazer'].isin(pronto_sel)

filtered = df[mask]


# ========================================================================
# HEADER
# ========================================================================
st.markdown("# 📦 Dashboard de Pedidos")
st.markdown(f"<p style='color: #888;'>Dados filtrados: {len(filtered):,} linhas de {len(df):,} | {filtered['Num. Pedido'].nunique()} PVs</p>", unsafe_allow_html=True)


# ========================================================================
# ABAS PRINCIPAIS
# ========================================================================
if fat_df is not None and len(fat_df) > 0:
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["📊 Visao Geral", "✅ Prontidao", "📅 Previsao Entrega", "📦 Estoque & SC/OP", "💰 Faturamento"])
else:
    tab1, tab2, tab3, tab4 = st.tabs(["📊 Visao Geral", "✅ Prontidao", "📅 Previsao Entrega", "📦 Estoque & SC/OP"])
    tab5 = None


# ===== ABA 1: VISAO GERAL =====
with tab1:
    # KPIs
    total_pvs = filtered['Num. Pedido'].nunique()
    total_linhas = len(filtered)
    em_aberto = len(filtered[filtered['Status_Pedido'] == 'EM ABERTO'])
    finalizados = len(filtered[filtered['Status_Pedido'] == 'FINALIZADO'])
    valor_total = filtered['Vlr.Total'].sum()
    valor_aberto = filtered.loc[filtered['Status_Pedido'] == 'EM ABERTO', 'Vlr.Total'].sum()
    pct = (finalizados / total_linhas * 100) if total_linhas > 0 else 0

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.markdown(kpi_card(f"{total_pvs}", "Total PVs"), unsafe_allow_html=True)
    c2.markdown(kpi_card(f"{total_linhas:,}", "Total Linhas"), unsafe_allow_html=True)
    c3.markdown(kpi_card(f"{em_aberto}", "Em Aberto", "alert"), unsafe_allow_html=True)
    c4.markdown(kpi_card(f"{finalizados}", "Finalizados", "ok"), unsafe_allow_html=True)
    c5.markdown(kpi_card(f"R$ {valor_aberto:,.0f}", "Valor Em Aberto", "warn"), unsafe_allow_html=True)
    c6.markdown(kpi_card(f"{pct:.1f}%", "% Conclusao", "ok" if pct > 80 else "warn"), unsafe_allow_html=True)

    st.markdown("")

    col1, col2 = st.columns(2)

    with col1:
        # Entrada de pedidos por mes
        if len(filtered) > 0:
            pedidos_mes = filtered.groupby('Mes_Emissao').size().reset_index(name='Qtd')
            pedidos_mes = pedidos_mes.sort_values('Mes_Emissao')
            fig = px.bar(pedidos_mes, x='Mes_Emissao', y='Qtd',
                        title='Entrada de Pedidos por Mes',
                        color_discrete_sequence=[CORES['azul_claro']])
            fig.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                font_color='#E0E0E0', xaxis_title='', yaxis_title='Quantidade',
                height=400
            )
            st.plotly_chart(fig, use_container_width=True)

    with col2:
        # Status
        if len(filtered) > 0:
            status_count = filtered['Status_Pedido'].value_counts().reset_index()
            status_count.columns = ['Status', 'Qtd']
            cores_status = {'FINALIZADO': CORES['verde'], 'EM ABERTO': CORES['vermelho']}
            fig2 = px.pie(status_count, values='Qtd', names='Status',
                         title='Distribuicao por Status',
                         color='Status', color_discrete_map=cores_status,
                         hole=0.4)
            fig2.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                font_color='#E0E0E0', height=400
            )
            st.plotly_chart(fig2, use_container_width=True)

    # Pedidos por vendedor
    if len(filtered) > 0 and 'Nome do Vendedor' in filtered.columns:
        vend_data = filtered.groupby(['Nome do Vendedor', 'Status_Pedido']).size().reset_index(name='Qtd')
        fig3 = px.bar(vend_data, x='Nome do Vendedor', y='Qtd', color='Status_Pedido',
                     title='Pedidos por Vendedor',
                     color_discrete_map={'FINALIZADO': CORES['verde'], 'EM ABERTO': CORES['vermelho']},
                     barmode='stack')
        fig3.update_layout(
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            font_color='#E0E0E0', xaxis_title='', yaxis_title='Quantidade',
            height=400
        )
        st.plotly_chart(fig3, use_container_width=True)


# ===== ABA 2: PRONTIDAO =====
with tab2:
    abertos = filtered[filtered['Status_Pedido'] == 'EM ABERTO']

    prontos = len(abertos[abertos['Pronto_para_Fazer'] == 'SIM'])
    parciais = len(abertos[abertos['Pronto_para_Fazer'].str.contains('PARCIAL', na=False)])
    nao_prontos = len(abertos[abertos['Pronto_para_Fazer'] == 'NAO'])
    erros = len(abertos[abertos['Acao_Necessaria'] == 'ERRO no CADASTRO'])

    # KPIs como botoes clicaveis
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        if st.button(f"✅ {prontos}\nProntos p/ Fazer", key="kpi_prontos", use_container_width=True):
            st.session_state['tab2_filtro'] = ('Pronto_para_Fazer', 'SIM')
    with c2:
        if st.button(f"⚠️ {parciais}\nParcialmente Prontos", key="kpi_parciais", use_container_width=True):
            st.session_state['tab2_filtro'] = ('Pronto_para_Fazer', 'PARCIAL')
    with c3:
        if st.button(f"🔴 {nao_prontos}\nNao Prontos", key="kpi_nao_prontos", use_container_width=True):
            st.session_state['tab2_filtro'] = ('Pronto_para_Fazer', 'NAO')
    with c4:
        if st.button(f"❌ {erros}\nERROS Cadastro", key="kpi_erros_t2", use_container_width=True):
            st.session_state['tab2_filtro'] = ('Acao_Necessaria', 'ERRO no CADASTRO')

    st.markdown("")

    col1, col2 = st.columns(2)

    filtro_pronto_sel = []
    filtro_tipo_sel = []

    with col1:
        if len(abertos) > 0:
            pronto_count = abertos['Pronto_para_Fazer'].value_counts().reset_index()
            pronto_count.columns = ['Status', 'Qtd']
            cores_pronto = {
                'SIM': CORES['verde'], 'NAO': CORES['vermelho'],
                'PARCIAL - Sem Follow-up': CORES['amarelo'],
                'PARCIAL - Sem Estoque': '#E67E22'
            }
            fig = px.pie(pronto_count, values='Qtd', names='Status',
                        title='Distribuicao - Pronto p/ Fazer? (clique para filtrar)',
                        color='Status', color_discrete_map=cores_pronto, hole=0.4)
            fig.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                font_color='#E0E0E0', height=400
            )
            evento_pronto = st.plotly_chart(fig, use_container_width=True, on_select="rerun", key="pie_pronto")

            if evento_pronto and evento_pronto.selection and evento_pronto.selection.points:
                filtro_pronto_sel = [p['label'] for p in evento_pronto.selection.points if 'label' in p]

    with col2:
        if len(abertos) > 0:
            tipo_count = abertos['Tipo_Produto'].value_counts().reset_index()
            tipo_count.columns = ['Tipo', 'Qtd']
            fig2 = px.bar(tipo_count, x='Tipo', y='Qtd',
                         title='Comprando vs Produzindo (clique para filtrar)',
                         color='Tipo',
                         color_discrete_map={'Comprando': CORES['azul_claro'], 'Produzindo': CORES['amarelo'],
                                           'Nao classificado': CORES['cinza']})
            fig2.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                font_color='#E0E0E0', height=400, showlegend=False
            )
            evento_tipo = st.plotly_chart(fig2, use_container_width=True, on_select="rerun", key="bar_tipo")

            if evento_tipo and evento_tipo.selection and evento_tipo.selection.points:
                filtro_tipo_sel = [p['x'] for p in evento_tipo.selection.points if 'x' in p]

    # Aplicar filtros: KPI button OU clique no grafico
    tabela_abertos = abertos.copy()
    filtro_info_partes = []

    # Filtro via KPI button
    kpi_filtro = st.session_state.get('tab2_filtro')
    if kpi_filtro:
        campo, valor = kpi_filtro
        if campo == 'Pronto_para_Fazer' and valor == 'PARCIAL':
            tabela_abertos = tabela_abertos[tabela_abertos[campo].str.contains('PARCIAL', na=False)]
            filtro_info_partes.append(f"Parcialmente Prontos")
        else:
            tabela_abertos = tabela_abertos[tabela_abertos[campo] == valor]
            filtro_info_partes.append(f"{valor}")

    # Filtro via clique no grafico (sobrepoe KPI se houver)
    if filtro_pronto_sel:
        tabela_abertos = abertos[abertos['Pronto_para_Fazer'].isin(filtro_pronto_sel)]
        filtro_info_partes = [f"Pronto: {', '.join(filtro_pronto_sel)}"]
        st.session_state.pop('tab2_filtro', None)
    if filtro_tipo_sel:
        tabela_abertos = tabela_abertos[tabela_abertos['Tipo_Produto'].isin(filtro_tipo_sel)]
        filtro_info_partes.append(f"Tipo: {', '.join(filtro_tipo_sel)}")
        st.session_state.pop('tab2_filtro', None)

    # Botao limpar filtro
    if filtro_info_partes or filtro_pronto_sel or filtro_tipo_sel or kpi_filtro:
        if st.button("🧹 Limpar filtros", key="limpar_t2"):
            st.session_state.pop('tab2_filtro', None)
            st.rerun()

    filtro_info = f" — Filtrado por: {' | '.join(filtro_info_partes)}" if filtro_info_partes else ""
    st.markdown(f"### 📋 Detalhes - Pedidos Em Aberto ({len(tabela_abertos)} itens){filtro_info}")

    if len(tabela_abertos) > 0:
        tab_cols = ['Num. Pedido', 'Item', 'Produto', 'Descricao do Produto', 'Tipo_Produto',
                   'Quantidade', 'Pronto_para_Fazer', 'Disponivel_Estoque', 'Acao_Necessaria',
                   'Prazo_Real_Entrega', 'Semana_Entrega', 'Ped Cliente']
        tab_cols = [c for c in tab_cols if c in tabela_abertos.columns]

        display_df = tabela_abertos[tab_cols].copy()
        display_df.columns = ['PV', 'Item', 'Produto', 'Descricao', 'Tipo',
                             'Qtd', 'Pronto?', 'Estoque', 'Acao',
                             'Prazo Entrega', 'Semana', 'Ped Cliente'][:len(tab_cols)]

        def color_pronto(val):
            if val == 'SIM': return 'background-color: #1a3a1a; color: #2ECC71'
            elif 'PARCIAL' in str(val): return 'background-color: #3a3a1a; color: #F39C12'
            elif val == 'NAO': return 'background-color: #3a1a1a; color: #E74C3C'
            return ''

        def color_acao(val):
            if val == 'ERRO no CADASTRO': return 'background-color: #C0392B; color: white; font-weight: bold'
            elif val == 'Estoque OK': return 'background-color: #1a3a1a; color: #2ECC71'
            elif val == 'Prazo a confirmar': return 'background-color: #3a3a1a; color: #F39C12'
            elif 'Necessario' in str(val): return 'background-color: #3a1a1a; color: #E74C3C'
            elif 'gerada' in str(val): return 'background-color: #3a3a1a; color: #F39C12'
            return ''

        styled = display_df.style.applymap(color_pronto, subset=['Pronto?'] if 'Pronto?' in display_df.columns else [])
        if 'Acao' in display_df.columns:
            styled = styled.applymap(color_acao, subset=['Acao'])

        st.dataframe(styled, use_container_width=True, height=500)
    else:
        st.info("Nenhum pedido em aberto com os filtros selecionados.")


# ===== ABA 3: PREVISAO ENTREGA =====
with tab3:
    abertos = filtered[filtered['Status_Pedido'] == 'EM ABERTO']

    atrasados_count = len(abertos[abertos['Dias_Atraso_Cliente'] > 0])
    no_prazo_count = len(abertos[(abertos['Dias_Atraso_Cliente'] <= 0) & (abertos['Dias_Atraso_Cliente'].notna())])
    sem_data_count = len(abertos[abertos['Dias_Atraso_Cliente'].isna()])
    media_atraso = abertos.loc[abertos['Dias_Atraso_Cliente'] > 0, 'Dias_Atraso_Cliente'].mean()

    # KPIs como botoes clicaveis
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        if st.button(f"🔴 {atrasados_count}\nAtrasados", key="kpi_atrasados_t3", use_container_width=True):
            st.session_state['tab3_filtro'] = 'atrasados'
    with c2:
        if st.button(f"✅ {no_prazo_count}\nNo Prazo", key="kpi_no_prazo_t3", use_container_width=True):
            st.session_state['tab3_filtro'] = 'no_prazo'
    with c3:
        if st.button(f"⚠️ {sem_data_count}\nSem Data", key="kpi_sem_data_t3", use_container_width=True):
            st.session_state['tab3_filtro'] = 'sem_data'
    with c4:
        st.markdown(kpi_card(f"{media_atraso:.0f} dias" if pd.notna(media_atraso) else "0", "Media Atraso", "warn"), unsafe_allow_html=True)

    st.markdown("")

    col1, col2 = st.columns(2)

    filtro_semana_sel = []

    with col1:
        semana_data = abertos[abertos['Semana_Entrega'].notna()]
        if len(semana_data) > 0:
            sem_count = semana_data['Semana_Entrega'].value_counts().reset_index()
            sem_count.columns = ['Semana', 'Qtd']
            sem_count = sem_count.sort_values('Semana')
            fig = px.bar(sem_count, x='Semana', y='Qtd',
                        title='Pedidos por Semana de Entrega (clique para filtrar)',
                        color_discrete_sequence=[CORES['azul_claro']])
            fig.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                font_color='#E0E0E0', height=400
            )
            evento_semana = st.plotly_chart(fig, use_container_width=True, on_select="rerun", key="bar_semana_t3")
            if evento_semana and evento_semana.selection and evento_semana.selection.points:
                filtro_semana_sel = [p['x'] for p in evento_semana.selection.points if 'x' in p]
        else:
            st.info("Nenhum pedido com semana de entrega definida.")

    with col2:
        top_atrasados = abertos.nlargest(10, 'Dias_Atraso_Cliente')[
            ['Num. Pedido', 'Produto', 'Dias_Atraso_Cliente']
        ].dropna(subset=['Dias_Atraso_Cliente'])

        if len(top_atrasados) > 0:
            top_atrasados['Label'] = top_atrasados['Num. Pedido'] + ' - ' + top_atrasados['Produto']
            fig2 = px.bar(top_atrasados, x='Dias_Atraso_Cliente', y='Label',
                         title='Top 10 - Mais Atrasados (vs Cliente)',
                         orientation='h', color_discrete_sequence=[CORES['vermelho']])
            fig2.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                font_color='#E0E0E0', yaxis_title='', xaxis_title='Dias de Atraso',
                height=400
            )
            st.plotly_chart(fig2, use_container_width=True)

    # Aplicar filtros
    tabela_t3 = abertos.copy()
    filtro_info_t3 = []

    kpi_filtro_t3 = st.session_state.get('tab3_filtro')
    if kpi_filtro_t3 == 'atrasados':
        tabela_t3 = tabela_t3[tabela_t3['Dias_Atraso_Cliente'] > 0]
        filtro_info_t3.append("Atrasados")
    elif kpi_filtro_t3 == 'no_prazo':
        tabela_t3 = tabela_t3[(tabela_t3['Dias_Atraso_Cliente'] <= 0) & (tabela_t3['Dias_Atraso_Cliente'].notna())]
        filtro_info_t3.append("No Prazo")
    elif kpi_filtro_t3 == 'sem_data':
        tabela_t3 = tabela_t3[tabela_t3['Dias_Atraso_Cliente'].isna()]
        filtro_info_t3.append("Sem Data")

    if filtro_semana_sel:
        tabela_t3 = tabela_t3[tabela_t3['Semana_Entrega'].isin(filtro_semana_sel)]
        filtro_info_t3 = [f"Semana: {', '.join(filtro_semana_sel)}"]
        st.session_state.pop('tab3_filtro', None)

    if filtro_info_t3 or kpi_filtro_t3:
        if st.button("🧹 Limpar filtros", key="limpar_t3"):
            st.session_state.pop('tab3_filtro', None)
            st.rerun()

    info_t3 = f" — Filtrado por: {' | '.join(filtro_info_t3)}" if filtro_info_t3 else ""
    st.markdown(f"### 📋 Previsao de Entrega ({len(tabela_t3)} itens){info_t3}")

    if len(tabela_t3) > 0:
        ent_cols = ['Num. Pedido', 'Item', 'Produto', 'Descricao do Produto', 'Quantidade',
                   'Ped Cliente', 'DT. Fat. Cli', 'DT. Ofertada', 'Prazo_Real_Entrega',
                   'Semana_Entrega', 'FU_Dt_Chegada_Autron', 'Dias_Atraso_Cliente', 'Pronto_para_Fazer']
        ent_cols = [c for c in ent_cols if c in tabela_t3.columns]

        ent_df = tabela_t3[ent_cols].copy().sort_values('Dias_Atraso_Cliente', ascending=False)
        rename_map = {
            'Num. Pedido': 'PV', 'Descricao do Produto': 'Descricao',
            'Quantidade': 'Qtd', 'Ped Cliente': 'Ped.Cliente',
            'DT. Fat. Cli': 'Dt.Solicitada', 'DT. Ofertada': 'Dt.Ofertada',
            'Prazo_Real_Entrega': 'Prazo Real', 'Semana_Entrega': 'Semana',
            'FU_Dt_Chegada_Autron': 'Chegada Autron',
            'Dias_Atraso_Cliente': 'Dias Atraso', 'Pronto_para_Fazer': 'Pronto?'
        }
        ent_df = ent_df.rename(columns=rename_map)

        def color_atraso(val):
            if pd.isna(val): return ''
            if val > 30: return 'background-color: #3a1a1a; color: #E74C3C'
            elif val > 0: return 'background-color: #3a3a1a; color: #F39C12'
            else: return 'background-color: #1a3a1a; color: #2ECC71'

        styled = ent_df.style
        if 'Dias Atraso' in ent_df.columns:
            styled = styled.applymap(color_atraso, subset=['Dias Atraso'])

        st.dataframe(styled, use_container_width=True, height=500)


# ===== ABA 4: ESTOQUE & SC/OP =====
with tab4:
    abertos = filtered[filtered['Status_Pedido'] == 'EM ABERTO']

    com_estoque = len(abertos[abertos['Disponivel_Estoque'] == 'SIM'])
    sem_estoque = len(abertos[abertos['Disponivel_Estoque'] == 'NAO'])
    parcial_est = len(abertos[abertos['Disponivel_Estoque'] == 'PARCIAL'])
    necessita_sc = len(abertos[abertos['Acao_Necessaria'] == 'Necessario gerar SC'])
    necessita_op = len(abertos[abertos['Acao_Necessaria'] == 'Necessario gerar OP'])
    erros = len(abertos[abertos['Acao_Necessaria'] == 'ERRO no CADASTRO'])

    # KPIs como botoes clicaveis
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    with c1:
        if st.button(f"✅ {com_estoque}\nCom Estoque", key="kpi_com_est", use_container_width=True):
            st.session_state['tab4_filtro'] = ('Disponivel_Estoque', 'SIM')
    with c2:
        if st.button(f"⚠️ {parcial_est}\nEstoque Parcial", key="kpi_parcial_est", use_container_width=True):
            st.session_state['tab4_filtro'] = ('Disponivel_Estoque', 'PARCIAL')
    with c3:
        if st.button(f"🔴 {sem_estoque}\nSem Estoque", key="kpi_sem_est", use_container_width=True):
            st.session_state['tab4_filtro'] = ('Disponivel_Estoque', 'NAO')
    with c4:
        if st.button(f"📋 {necessita_sc}\nNecessitam SC", key="kpi_nec_sc", use_container_width=True):
            st.session_state['tab4_filtro'] = ('Acao_Necessaria', 'Necessario gerar SC')
    with c5:
        if st.button(f"🏭 {necessita_op}\nNecessitam OP", key="kpi_nec_op", use_container_width=True):
            st.session_state['tab4_filtro'] = ('Acao_Necessaria', 'Necessario gerar OP')
    with c6:
        if st.button(f"❌ {erros}\nERROS Cadastro", key="kpi_erros_t4", use_container_width=True):
            st.session_state['tab4_filtro'] = ('Acao_Necessaria', 'ERRO no CADASTRO')

    st.markdown("")

    col1, col2 = st.columns(2)

    filtro_est_sel = []
    filtro_acao_sel = []

    with col1:
        if len(abertos) > 0:
            est_count = abertos['Disponivel_Estoque'].value_counts().reset_index()
            est_count.columns = ['Disponibilidade', 'Qtd']
            cores_est = {'SIM': CORES['verde'], 'NAO': CORES['vermelho'], 'PARCIAL': CORES['amarelo'], 'Serviço': CORES['amarelo']}
            fig = px.pie(est_count, values='Qtd', names='Disponibilidade',
                        title='Disponibilidade de Estoque (clique para filtrar)',
                        color='Disponibilidade', color_discrete_map=cores_est, hole=0.4)
            fig.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                font_color='#E0E0E0', height=380
            )
            evento_est = st.plotly_chart(fig, use_container_width=True, on_select="rerun", key="pie_estoque_t4")
            if evento_est and evento_est.selection and evento_est.selection.points:
                filtro_est_sel = [p['label'] for p in evento_est.selection.points if 'label' in p]

    with col2:
        if len(abertos) > 0:
            acao_count = abertos[abertos['Acao_Necessaria'] != 'Finalizado']['Acao_Necessaria'].value_counts().reset_index()
            acao_count.columns = ['Acao', 'Qtd']
            cores_acao = {
                'Estoque OK': CORES['verde'],
                'Necessario gerar SC': CORES['vermelho'],
                'Necessario gerar OP': '#E67E22',
                'SC gerada - Aguardando': CORES['amarelo'],
                'OP gerada - Aguardando': '#F1C40F',
                'ERRO no CADASTRO': CORES['vermelho_escuro'],
                'Prazo a confirmar': CORES['amarelo'],
                'Verificar classificacao': CORES['cinza']
            }
            fig2 = px.bar(acao_count, x='Acao', y='Qtd',
                         title='Acoes Necessarias (clique para filtrar)',
                         color='Acao', color_discrete_map=cores_acao)
            fig2.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                font_color='#E0E0E0', height=380, showlegend=False,
                xaxis_title='', yaxis_title='Quantidade'
            )
            evento_acao = st.plotly_chart(fig2, use_container_width=True, on_select="rerun", key="bar_acao_t4")
            if evento_acao and evento_acao.selection and evento_acao.selection.points:
                filtro_acao_sel = [p['x'] for p in evento_acao.selection.points if 'x' in p]

    # Aplicar filtros: KPI button OU clique no grafico
    tabela_t4 = abertos.copy()
    filtro_info_t4 = []

    # Filtro via KPI button
    kpi_filtro_t4 = st.session_state.get('tab4_filtro')
    if kpi_filtro_t4:
        campo, valor = kpi_filtro_t4
        tabela_t4 = tabela_t4[tabela_t4[campo] == valor]
        filtro_info_t4.append(f"{valor}")

    # Filtro via clique no grafico (sobrepoe KPI)
    if filtro_est_sel:
        tabela_t4 = abertos[abertos['Disponivel_Estoque'].isin(filtro_est_sel)]
        filtro_info_t4 = [f"Estoque: {', '.join(filtro_est_sel)}"]
        st.session_state.pop('tab4_filtro', None)
    if filtro_acao_sel:
        tabela_t4 = tabela_t4[tabela_t4['Acao_Necessaria'].isin(filtro_acao_sel)]
        filtro_info_t4.append(f"Acao: {', '.join(filtro_acao_sel)}")
        st.session_state.pop('tab4_filtro', None)

    if filtro_info_t4 or kpi_filtro_t4:
        if st.button("🧹 Limpar filtros", key="limpar_t4"):
            st.session_state.pop('tab4_filtro', None)
            st.rerun()

    info_t4 = f" — Filtrado por: {' | '.join(filtro_info_t4)}" if filtro_info_t4 else ""
    st.markdown(f"### 📋 Detalhe Estoque & SC/OP ({len(tabela_t4)} itens){info_t4}")

    if len(tabela_t4) > 0:
        est_cols = ['Num. Pedido', 'Item', 'Produto', 'Descricao do Produto', 'Tipo_Produto',
                   'Quantidade', 'Estoque_Disponivel', 'Qtd_Alocada', 'Disponivel_Estoque',
                   'Acao_Necessaria', 'Numero SC', 'Numero OP', 'FU_OP_na_SC', 'DT Emissao']
        est_cols = [c for c in est_cols if c in tabela_t4.columns]

        est_df = tabela_t4[est_cols].copy().sort_values(['Produto', 'DT Emissao'])
        rename_est = {
            'Num. Pedido': 'PV', 'Descricao do Produto': 'Descricao',
            'Tipo_Produto': 'Tipo', 'Quantidade': 'Qtd Pedida',
            'Estoque_Disponivel': 'Estoque', 'Qtd_Alocada': 'Alocado',
            'Disponivel_Estoque': 'Disponivel', 'Acao_Necessaria': 'Acao',
            'Numero SC': 'SC', 'Numero OP': 'OP', 'FU_OP_na_SC': 'OP na SC',
            'DT Emissao': 'Dt.Emissao'
        }
        est_df = est_df.rename(columns=rename_est)

        def color_acao_tab(val):
            if val == 'ERRO no CADASTRO': return 'background-color: #C0392B; color: white; font-weight: bold'
            elif val == 'Estoque OK': return 'background-color: #1a3a1a; color: #2ECC71'
            elif val == 'Prazo a confirmar': return 'background-color: #3a3a1a; color: #F39C12'
            elif 'Necessario' in str(val): return 'background-color: #3a1a1a; color: #E74C3C'
            elif 'gerada' in str(val): return 'background-color: #3a3a1a; color: #F39C12'
            return ''

        styled = est_df.style
        if 'Acao' in est_df.columns:
            styled = styled.applymap(color_acao_tab, subset=['Acao'])

        st.dataframe(styled, use_container_width=True, height=500)

    # Alertas de erro
    erros_df = tabela_t4[tabela_t4['Acao_Necessaria'] == 'ERRO no CADASTRO'] if len(tabela_t4) > 0 else pd.DataFrame()
    if len(erros_df) > 0:
        st.markdown("### ⚠️ ERROS DE CADASTRO - Item Comprando com OP")
        st.error(f"Foram encontrados {len(erros_df)} itens classificados como 'Comprando' que possuem OP gerada. Verificar cadastro!")

        erro_display = erros_df[['Num. Pedido', 'Item', 'Produto', 'Descricao do Produto', 'Numero OP', 'FU_OP_na_SC']].copy()
        erro_display.columns = ['PV', 'Item', 'Produto', 'Descricao', 'OP (Entrada)', 'OP na SC (Follow-up)']
        st.dataframe(erro_display, use_container_width=True)


# ===== ABA 5: FATURAMENTO =====
if tab5 is not None:
    with tab5:
        fat = fat_df.copy()
        fat['Emissao'] = pd.to_datetime(fat['Emissao'], errors='coerce')
        fat['Faturamento Bruto'] = pd.to_numeric(fat.get('Faturamento Bruto'), errors='coerce')
        fat['Faturamento Liquido'] = pd.to_numeric(fat.get('Faturamento Liquido'), errors='coerce')
        fat['Margem Liquida (R$)'] = pd.to_numeric(fat.get('Margem Liquida (R$)'), errors='coerce')
        fat['Margem Liquida (%)'] = pd.to_numeric(fat.get('Margem Liquida (%) por NF Faturada'), errors='coerce')
        fat['Mes_Fat'] = fat['Emissao'].dt.to_period('M').astype(str)

        fat_bruto = fat['Faturamento Bruto'].sum()
        fat_liq = fat['Faturamento Liquido'].sum()
        fat_margem_pct = fat['Margem Liquida (%)'].mean()
        fat_nfs = fat['Num. Docto.'].nunique() if 'Num. Docto.' in fat.columns else 0

        c1, c2, c3, c4 = st.columns(4)
        c1.markdown(kpi_card(f"R$ {fat_bruto:,.0f}", "Faturamento Bruto"), unsafe_allow_html=True)
        c2.markdown(kpi_card(f"R$ {fat_liq:,.0f}", "Faturamento Liquido", "ok"), unsafe_allow_html=True)
        c3.markdown(kpi_card(f"{fat_margem_pct:.1f}%", "Margem Liquida Media", "ok" if fat_margem_pct >= 30 else "warn"), unsafe_allow_html=True)
        c4.markdown(kpi_card(f"{fat_nfs}", "Notas Fiscais"), unsafe_allow_html=True)

        st.markdown("")

        col1, col2 = st.columns(2)

        filtro_mes_fat = []
        filtro_vend_fat = []

        with col1:
            fat_mes = fat.groupby('Mes_Fat').agg({'Faturamento Liquido': 'sum'}).reset_index()
            fat_mes = fat_mes.sort_values('Mes_Fat')
            fig = px.bar(fat_mes, x='Mes_Fat', y='Faturamento Liquido',
                        title='Faturamento Liquido por Mes (clique para filtrar)',
                        color_discrete_sequence=[CORES['verde']])
            fig.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                font_color='#E0E0E0', xaxis_title='', yaxis_title='R$', height=400
            )
            evento_mes_fat = st.plotly_chart(fig, use_container_width=True, on_select="rerun", key="bar_mes_fat")
            if evento_mes_fat and evento_mes_fat.selection and evento_mes_fat.selection.points:
                filtro_mes_fat = [p['x'] for p in evento_mes_fat.selection.points if 'x' in p]

        with col2:
            if 'Nome do Vendedor' in fat.columns:
                fat_vend = fat.groupby('Nome do Vendedor').agg(
                    {'Faturamento Liquido': 'sum'}
                ).reset_index().sort_values('Faturamento Liquido', ascending=True).tail(10)
                fig2 = px.bar(fat_vend, x='Faturamento Liquido', y='Nome do Vendedor',
                             title='Top 10 Vendedores (clique para filtrar)',
                             orientation='h', color_discrete_sequence=[CORES['azul_claro']])
                fig2.update_layout(
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    font_color='#E0E0E0', yaxis_title='', xaxis_title='R$', height=400
                )
                evento_vend = st.plotly_chart(fig2, use_container_width=True, on_select="rerun", key="bar_vend_fat")
                if evento_vend and evento_vend.selection and evento_vend.selection.points:
                    filtro_vend_fat = [p['y'] for p in evento_vend.selection.points if 'y' in p]

        # Aplicar filtros
        tabela_fat = fat.copy()
        filtro_info_fat = []

        if filtro_vend_fat:
            tabela_fat = tabela_fat[tabela_fat['Nome do Vendedor'].isin(filtro_vend_fat)]
            filtro_info_fat.append(f"Vendedor: {', '.join(filtro_vend_fat)}")
        if filtro_mes_fat:
            tabela_fat = tabela_fat[tabela_fat['Mes_Fat'].isin(filtro_mes_fat)]
            filtro_info_fat.append(f"Mes: {', '.join(filtro_mes_fat)}")

        if filtro_info_fat:
            # KPIs atualizados para o filtro
            fat_bruto_f = tabela_fat['Faturamento Bruto'].sum()
            fat_liq_f = tabela_fat['Faturamento Liquido'].sum()
            fat_margem_f = tabela_fat['Margem Liquida (%)'].mean()
            fat_nfs_f = tabela_fat['Num. Docto.'].nunique() if 'Num. Docto.' in tabela_fat.columns else 0

            st.markdown("#### 📊 Resumo do filtro aplicado")
            r1, r2, r3, r4 = st.columns(4)
            r1.markdown(kpi_card(f"R$ {fat_bruto_f:,.0f}", "Fat. Bruto (filtrado)"), unsafe_allow_html=True)
            r2.markdown(kpi_card(f"R$ {fat_liq_f:,.0f}", "Fat. Liquido (filtrado)", "ok"), unsafe_allow_html=True)
            r3.markdown(kpi_card(f"{fat_margem_f:.1f}%" if pd.notna(fat_margem_f) else "N/A", "Margem Media (filtrado)", "ok" if pd.notna(fat_margem_f) and fat_margem_f >= 30 else "warn"), unsafe_allow_html=True)
            r4.markdown(kpi_card(f"{fat_nfs_f}", "NFs (filtrado)"), unsafe_allow_html=True)

            if st.button("🧹 Limpar filtros", key="limpar_t5"):
                st.rerun()

        info_fat = f" — Filtrado por: {' | '.join(filtro_info_fat)}" if filtro_info_fat else ""
        st.markdown(f"### 📋 Detalhe do Faturamento ({len(tabela_fat)} itens){info_fat}")

        fat_display_cols = [c for c in ['Emissao', 'Num. Docto.', 'No do Pedido', 'Produto',
            'Descricao Produto', 'Quantidade', 'Razao Social', 'Nome Fantasia', 'UF',
            'Faturamento Bruto', 'Faturamento Liquido', 'Margem Liquida (R$)',
            'Margem Liquida (%) por NF Faturada',
            'Nome do Vendedor', 'Tipo Negocio'] if c in tabela_fat.columns]
        fat_table = tabela_fat[fat_display_cols].copy().sort_values('Emissao', ascending=False)
        fat_table = fat_table.rename(columns={'Margem Liquida (%) por NF Faturada': 'Margem Liquida (%)'})
        st.dataframe(fat_table, use_container_width=True, height=500)
